"""vote URL Configuration

The `urlpatterns` list routes URLs to views. For more information please see:
    https://docs.djangoproject.com/en/2.2/topics/http/urls/
Examples:
Function views
    1. Add an import:  from my_app import views
    2. Add a URL to urlpatterns:  path('', views.home, name='home')
Class-based views
    1. Add an import:  from other_app.views import Home
    2. Add a URL to urlpatterns:  path('', Home.as_view(), name='home')
Including another URLconf
    1. Import the include() function: from django.urls import include, path
    2. Add a URL to urlpatterns:  path('blog/', include('blog.urls'))
"""
import subprocess
from django.utils import timezone
from django.contrib import admin
from django.urls import path
from django.views.decorators.csrf import csrf_exempt
from django.views.generic.base import TemplateView
from django.conf import settings
from django.shortcuts import redirect
from django.conf.urls.static import static
import json
from django.http import JsonResponse
from django.http import HttpResponseForbidden
from django.contrib.auth import authenticate, login
# from django.contrib.auth.decorators import login_required
from openpyxl import load_workbook
from background_task import background
from .models import Works, UserVoteLog, UploadStatus
import os
# import the logging library
import logging
# from weixin.client import WeixinMpAPI
from weixin.login import WeixinLogin

# Get an instance of a logger
logger = logging.getLogger(__name__)

# deploy_domain = '142.93.185.148:8888'
# deploy_domain = '192.168.1.103:8080'
deploy_domain = 'vote.ilingyue.cn'
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
wechat_appid = 'wx1f8565a373ffdead'
wechat_appsecret = '8e16759000304ad924851a304aea4efa'
wx_login = WeixinLogin(wechat_appid, wechat_appsecret)


@csrf_exempt
def wechat_login(request):
    # if request.session.get('openid', None):
    #     return redirect('https://vote.ilingyue.cn/index.html')
    # else:
    return redirect(wx_login.authorize(
        'https://vote.ilingyue.cn/authorized',
        'snsapi_userinfo'
    ))


@csrf_exempt
def wechat_authorize(request):
    if request.method == 'GET':
        code = request.GET.get('code', None)
        if code:
            data = wx_login.access_token(code)
            request.session['openid'] = data.openid
            return redirect('https://vote.ilingyue.cn/index.html?openid=' +
                            data.openid)
        else:
            return HttpResponseForbidden()
    else:
        return HttpResponseForbidden()


@csrf_exempt
def admin_login(request):
    if request.method == 'POST':
        data = json.loads(str(request.body, encoding='utf8'))
        username = data.get('username', '')
        password = data.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return JsonResponse({
                'code': 'success'
            })
        else:
            return JsonResponse({'code': 'error'})
    else:
        return HttpResponseForbidden()


def generate_pdf(url, tag):
    print('开始处理 ', tag, ': ', url)
    subprocess.run(['bash', os.path.join(BASE_DIR,
                                         'weixinCapture',
                                         'generate_pdf.sh'),
                    url,
                    str(tag)])
    print('完成处理 ', tag, ': ', url)
    return 'http://' + deploy_domain + '/media/' + str(tag) + '.pdf'


@background(schedule=0)
def process_works(max_votes, filename):
    try:
        works = {}
        idx = 2
        wb = load_workbook(filename=filename)
        sheet = wb['Sheet1']
        if not UploadStatus.objects.exists():
            status = UploadStatus()
            status.save()
        status = UploadStatus.objects.all()[0]
        status.status = 'loading'
        status.save()
        while True:
            tag = sheet['A' + str(idx)].value
            if tag is not None:
                works[tag] = {
                    'tag': tag,
                    'title': sheet['B' + str(idx)].value,
                    'desc': sheet['C' + str(idx)].value,
                    'imageURL': sheet['D' + str(idx)].value,
                    'origin_url': sheet['E' + str(idx)].value,
                    'pdf_url': generate_pdf(sheet['E' + str(idx)].value, tag),
                    'votes': 0
                }
                idx += 1
            else:
                break
        if len(Works.objects.all()) > 0:
            Works.objects.all()[0].delete()
        w = Works(works=json.dumps(works))
        w.max_votes = max_votes
        w.save()
        status = UploadStatus.objects.all()[0]
        status.status = 'free'
        status.save()
    except Exception as e:
        logger.error(e)


# @login_required
@csrf_exempt
def upload(request):
    if request.method == 'POST':
        # data = json.loads(str(request.body, encoding='utf8'))
        data = request.POST
        max_votes = data.get('max_votes', 1)
        works = request.FILES
        works = works.getlist(list(works.keys())[0])[0]
        # print(works)
        # print(json.dumps(works))
        with open('temp_works.xlsx', 'wb+') as f:
            for chunk in works.chunks():
                f.write(chunk)
        wb = load_workbook(filename='temp_works.xlsx')
        sheet = wb['Sheet1']
        valid = all([
            sheet['A1'].value == '作品 id',
            sheet['B1'].value == '作品名称',
            sheet['C1'].value == '作品简介',
            sheet['D1'].value == '图片简介',
            sheet['E1'].value == '作品链接'
        ])
        cnt = 0
        while True:
            if sheet['A' + str(cnt + 2)].value is not None:
                cnt += 1
            else:
                break
        if valid:
            process_works(max_votes, 'temp_works.xlsx',
                          schedule=timezone.now())
            return JsonResponse({'code': 'success', 'data': cnt})
        else:
            return JsonResponse({
                'code': 'error',
                'data': '文件格式错误, 必须按照模板文件,!'
            })
    else:
        return HttpResponseForbidden()


@csrf_exempt
# @login_required
def get_results(request):
    if request.method == 'POST':
        status = UploadStatus.objects
        if status.exists() and status.all()[0].status == 'loading':
            return JsonResponse({
                'code': 'error',
                'data': '作品正在后台生成中!'
            })
        elif len(Works.objects.all()) > 0:
            works = json.loads(Works.objects.all()[0].works)
            return JsonResponse({
                'code': 'success',
                'data': works
            })
        else:
            return JsonResponse({'code': 'error', 'data': '还未上传作品列表!'})
    else:
        return HttpResponseForbidden()


@csrf_exempt
def vote(req):
    if req.method == 'POST':
        data = json.loads(str(req.body, encoding='utf8'))
        user_id = data.get('user_id', None)
        tag = data.get('tag', None)
        if user_id and tag:
            try:
                if UserVoteLog.objects.exists():
                    u = UserVoteLog.objects.filter(wechat_openid=user_id)
                else:
                    u = []
                max_votes = Works.objects.all()[0].max_votes
                if len(u) >= max_votes:
                    return JsonResponse({'code': 'error', 'data': '您已经投过了!'})
                else:
                    log = UserVoteLog()
                    log.work_tag = tag
                    log.wechat_openid = user_id
                    log.save()
                    works = Works.objects.all()[0]
                    works_obj = json.loads(works.works)
                    works_obj[str(tag)]['votes'] += 1
                    works.works = json.dumps(works_obj)
                    works.save()
                    return JsonResponse({'code': 'success'})
            except Exception as e:
                return JsonResponse({
                    'code': 'error',
                    'data': str(e)
                })
        else:
            return JsonResponse({'code': 'error',
                                 'data': 'user_id={}, tag={}'.format(str(user_id), # noqa
                                                                     str(tag))}) # noqa
    else:
        return HttpResponseForbidden()


@csrf_exempt
def get_works(req):
    if req.method == 'POST':
        if UploadStatus.objects.exists():
            if UploadStatus.objects.all()[0].status == 'loading': # noqa
                return JsonResponse({
                    'code': 'error',
                    'data': '作品正在后台生成中'
                })
        works = Works.objects.all()
        if len(works) > 0:
            data = json.loads(str(req.body, encoding='utf8'))
            user_id = data.get('userId', None)
            if user_id:
                works = works[0]
                max_votes = works.max_votes
                works = json.loads(works.works)
                if UserVoteLog.objects.exists():
                    u = UserVoteLog.objects.filter(wechat_openid=user_id)
                    # for i in UserVoteLog.objects.all():
                    #     print(i.wechat_openid, i.work_tag)
                    for work in works:
                        works[work]['voted'] = False
                    if len(u) > 0:
                        for vote in u:
                            works[vote.work_tag]['voted'] = True
                return JsonResponse({
                    'code': 'success',
                    'data': {
                        'works': works,
                        'max_votes': max_votes
                    }
                })
            else:
                return JsonResponse({
                    'code': 'error'
                })
        else:
            return JsonResponse({
                'code': 'error',
                'data': '当前没有作品可供投票'
            })
    else:
        return HttpResponseForbidden()


urlpatterns = [
    path('admin/', admin.site.urls),
    path('adminLogin', admin_login),
    path('get_results', get_results),
    path('vote', vote),
    path('getWorks', get_works),
    path('upload', upload),
    path('wechat_login', wechat_login),
    path('authorized', wechat_authorize),
    path('login', wechat_login),
    path('index.html', TemplateView.as_view(
        template_name="index.html")),
] + static(settings.MEDIA_URL, document_root=settings.MEDIA_ROOT)
